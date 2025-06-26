from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
import importlib
import pandas as pd
from tkinter import messagebox
from .config import get_config, DOSSIER_SORTIE
from .data_processing import charger_fichier_excel, extraire_date, filtrer_donnees
from .utils import configurer_impression, configurer_styles

def generer_rapport_combine(chemins_fichiers, root):
    try:
        config = get_config()
        if not chemins_fichiers:
            messagebox.showerror("Erreur", "Aucun fichier sélectionné.")
            return None

        wb = Workbook()
        styles = configurer_styles()
        date_rapport = None

        for fichier_type, chemin_fichier in chemins_fichiers.items():
            if fichier_type not in config['fichiers']:
                messagebox.showwarning("Avertissement", f"Type de fichier inconnu : {fichier_type}")
                continue

            df = charger_fichier_excel(chemin_fichier, config['fichiers'][fichier_type]['feuille'])
            if df is None:
                continue

            date_rapport = extraire_date(chemin_fichier, df)
            if date_rapport is None:
                continue

            for rapport_type, rapport_info in config['fichiers'][fichier_type]['rapports'].items():
                df_rapport = filtrer_donnees(df, fichier_type, rapport_type)
                if df_rapport is not None and not df_rapport.empty:
                    try:
                        filter_module = importlib.import_module(f".filters.{rapport_type}", package=__package__)
                        generer_rapport_feuille = getattr(filter_module, f"generer_rapport_{rapport_type}_feuille")
                        ws = wb.create_sheet(f"{rapport_info['feuille']}")
                        if rapport_info.get('dialog', False):
                            dialog_class = getattr(filter_module, "MultiBennesDialog", None)
                            if dialog_class:
                                df_multibenne = df[df['Genre'].str.upper().str.strip() == 'MULTIBENNE']
                                multibennes_data = {
                                    f"MULTI BENNES N°{i+1}": {
                                        'Matricule': row['Matricule'],
                                        'Num de parc': row['Num de parc'],
                                        'Secteurs': row['Secteur'].split(', ') if isinstance(row['Secteur'], str) else [row['Secteur']]
                                    } for i, (_, row) in enumerate(df_multibenne.iterrows())
                                }
                                dialog = dialog_class(root, multibennes_data)
                                root.wait_window(dialog)
                                if dialog.result is None:
                                    messagebox.showinfo("Information", f"Génération du rapport {rapport_info['feuille']} annulée par l'utilisateur.")
                                    continue
                                generer_rapport_feuille(ws, df_rapport, date_rapport, dialog.result, multibennes_data, styles)
                            else:
                                messagebox.showerror("Erreur", f"Boîte de dialogue non définie pour {rapport_type}")
                                continue
                        else:
                            generer_rapport_feuille(ws, df_rapport, date_rapport, styles)
                        configurer_impression(ws)
                    except (ImportError, AttributeError) as e:
                        messagebox.showerror("Erreur", f"Erreur lors du chargement du module {rapport_type} : {e}")
                        continue
                else:
                    print(f"Aucune donnée pour {rapport_info['feuille']} dans {os.path.basename(chemin_fichier)}")

        if date_rapport is None:
            messagebox.showerror("Erreur", "Aucune date valide trouvée dans les fichiers.")
            return None

        nom_fichier_sortie = f"Rapport Journalier Combiné {date_rapport.strftime('%d-%m-%Y')}.xlsx"
        chemin_sortie = os.path.join(DOSSIER_SORTIE, nom_fichier_sortie)

        try:
            wb.save(chemin_sortie)
            return nom_fichier_sortie
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement du rapport : {e}")
            return None
    except Exception as e:
        messagebox.showerror("Erreur", f"Erreur inattendue lors de la génération du rapport combiné : {e}")
        return None

def generer_rapport_individuel(fichier_type, chemin_fichier, rapport_type, root=None):
    try:
        config = get_config()
        if fichier_type not in config['fichiers']:
            messagebox.showerror("Erreur", f"Type de fichier inconnu : {fichier_type}")
            return None

        df = charger_fichier_excel(chemin_fichier, config['fichiers'][fichier_type]['feuille'])
        if df is None:
            return None

        date_rapport = extraire_date(chemin_fichier, df)
        if date_rapport is None:
            return None

        df_rapport = filtrer_donnees(df, fichier_type, rapport_type)
        if df_rapport is None or df_rapport.empty:
            messagebox.showwarning("Avertissement", f"Aucune donnée pour {config['fichiers'][fichier_type]['rapports'][rapport_type]['feuille']} dans {os.path.basename(chemin_fichier)}")
            return None

        nom_fichier_sortie = f"Rapport Journalier {config['fichiers'][fichier_type]['rapports'][rapport_type]['feuille'].upper()} {date_rapport.strftime('%d-%m-%Y')}.xlsx"
        chemin_sortie = os.path.join(DOSSIER_SORTIE, nom_fichier_sortie)
        wb = Workbook()
        ws = wb.active
        ws.title = config['fichiers'][fichier_type]['rapports'][rapport_type]['feuille']
        styles = configurer_styles()

        try:
            filter_module = importlib.import_module(f".filters.{rapport_type}", package=__package__)
            generer_rapport_feuille = getattr(filter_module, f"generer_rapport_{rapport_type}_feuille")
            if config['fichiers'][fichier_type]['rapports'][rapport_type].get('dialog', False):
                dialog_class = getattr(filter_module, "MultiBennesDialog", None)
                if dialog_class:
                    df_multibenne = df[df['Genre'].str.upper().str.strip() == 'MULTIBENNE']
                    multibennes_data = {
                        f"MULTI BENNES N°{i+1}": {
                            'Matricule': row['Matricule'],
                            'Num de parc': row['Num de parc'],
                            'Secteurs': row['Secteur'].split(', ') if isinstance(row['Secteur'], str) else [row['Secteur']]
                        } for i, (_, row) in enumerate(df_multibenne.iterrows())
                    }
                    dialog = dialog_class(root, multibennes_data)
                    root.wait_window(dialog)
                    if dialog.result is None:
                        messagebox.showinfo("Information", f"Génération du rapport {config['fichiers'][fichier_type]['rapports'][rapport_type]['feuille']} annulée par l'utilisateur.")
                        return None
                    generer_rapport_feuille(ws, df_rapport, date_rapport, dialog.result, multibennes_data, styles)
                else:
                    messagebox.showerror("Erreur", f"Boîte de dialogue non définie pour {rapport_type}")
                    return None
            else:
                generer_rapport_feuille(ws, df_rapport, date_rapport, styles)
        except (ImportError, AttributeError) as e:
            messagebox.showerror("Erreur", f"Erreur lors du chargement du module {rapport_type} : {e}")
            return None

        try:
            wb.save(chemin_sortie)
            return nom_fichier_sortie
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement du rapport : {e}")
            return None
    except Exception as e:
        messagebox.showerror("Erreur", f"Erreur inattendue lors de la génération du rapport {rapport_type} : {e}")
        return None