from openpyxl.utils import get_column_letter
from datetime import datetime

def generer_rapport_{type}_feuille(ws, df, date_rapport{dialog_param}, styles):
    """
    Template pour générer une feuille de rapport pour {type}.
    Placeholders: {type}, {nom_feuille}, {dialog_param}.
    {dialog_param} sera remplacé par ', dialog_result, multibennes_data' pour multibenne, ou '' pour les autres.
    """
    ws.append([f"Rapport {nom_feuille}", date_rapport.strftime("%d/%m/%Y")])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1).font = styles['entete_font']
    ws.cell(row=1, column=1).alignment = styles['entete_alignment']

    headers = ["Zone", "Date d'Affectation", "Shift", "Secteur", "Equipment", "Genre", "Num de parc", "Matricule"]
    ws.append(headers)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col)
        cell.font = styles['entete_font']
        cell.alignment = styles['entete_alignment']
        cell.border = styles['border']

    row = 3
    for _, record in df.iterrows():
        ws.append([
            record.get('Zone', ''),
            date_rapport.strftime("%d/%m/%Y"),
            record.get('Shift', ''),
            record.get('Secteur', ''),
            record.get('Equipment', ''),
            record.get('Genre', ''),
            record.get('Num de parc', ''),
            record.get('Matricule', '')
        ])
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.alignment = styles['cell_alignment']
            cell.border = styles['border']
        row += 1

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15